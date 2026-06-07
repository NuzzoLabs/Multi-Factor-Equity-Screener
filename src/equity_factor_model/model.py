# -*- coding: utf-8 -*-
"""
Equity factor model core logic.

Research-code notice: this module was converted from a personal analysis script.
Some methods may require additional validation before investment or production use.

Created on Mon Oct 20 17:55:44 2025

V1.0: Automated equity screener built on Alpha Vantage fundamental data.

High-level flow:
1. Load securities (tickers)
2. Collect raw financial statements + metadata
3. Compute standardized factor scores:
   - Liquidity
   - Growth
   - Capital Efficiency
   - Coverage
   - Value
   - Risk
4. Aggregate into a client-agnostic factor DataFrame

Design philosophy:
- Quantitative fundamental analysis inspired by Warren Buffett / quality-first investing
- Emphasis on capital efficiency, durability, reinvestment quality, and downside protection
- No price prediction or short-term trading logic

@author: nnuzz_rpjwflj
"""

import requests
import numpy as np
import pandas as pd
from datetime import date, datetime, timedelta
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Display all rows/columns when printing DataFrames (debugging & inspection)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)


# ================================
# Alpha Vantage Client Wrapper
# ================================
class AlphaVantageClient:
    """
    Wrapper class for Alpha Vantage API calls and downstream financial computations.
    Centralizes:
    - API interaction
    - Data sanitation
    - Factor construction
    """

    # Free API Key (25 requests/day)
    # Premium API Key (75 requests/min)
    def __init__(self, api_key):
        self.api_key = api_key  # Store API key once per client instance

    # The following function is for calculating z-scores
    def calculate_z(self, X, X_mean, X_sd):
        return (X - X_mean) / X_sd
        
    # The following function is for calculating weights based on macro-economic factors
    def multiplier(self, z, relation, cap=3.0, max_tilt=1.5):
        """
        Compute macro-adjusted multiplier for a factor based on z-score and relationship.

        Parameters:
        - z: macro z-score
        - relation: 'direct' or 'inverse'
            direct -> higher z increases weight
            inverse -> higher z increases weight for defensive factors
        - cap: maximum absolute z-score considered
        - max_tilt: max multiplier applied
        """
        # Cap extremes
        z = max(min(z, cap), -cap)
        
        # Base linear tilt
        tilt = (z / cap) * (max_tilt - 1)

        # Direct vs inverse logic
        if relation == 'direct':
            multiplier = 1 - tilt          # +z → lower weight
        elif relation == 'inverse':
            multiplier = 1 + tilt          # +z → higher weight
        else:
            raise ValueError("relation must be 'direct' or 'inverse'")

        return multiplier

    # ----------------------------
    # Safe JSON extraction helpers
    # ----------------------------
    def safe_get_report(self, report, key):
        """
        Safely extract a numeric value from a dict-based Alpha Vantage report.
        Returns NaN if:
        - key is missing
        - value is None or string "None"
        - conversion fails
        """
        try:
            val = report.get(key, np.nan)
            return float(val) if val not in [None, "None"] else np.nan
        except Exception:
            return np.nan        

    def safe_get_list(self, report, list_key, field):
        """
        Safely extract a list of values from list-based Alpha Vantage JSON structures
        (e.g., insider transactions, shares outstanding history).

        Returns empty list if:
        - key missing
        - malformed data
        """
        try:
            data = report.get(list_key, [])
            if not isinstance(data, list):
                return []
            return [
                item.get(field)
                for item in data
                if item.get(field) not in [None, "None"]
            ]
        except Exception:
            return []

    # ----------------------------
    # Low-level API fetch
    # ----------------------------
    def _fetch_json(self, function, tickers):
        """
        Fetch raw JSON responses from Alpha Vantage for a given function
        (OVERVIEW, INCOME_STATEMENT, etc.) for all tickers.
        """
        urls = [
            f'https://www.alphavantage.co/query?function={function}&symbol={ticker}&apikey={self.api_key}'
            for ticker in tickers
        ]
        responses = [requests.get(url) for url in urls]
        return [resp.json() for resp in responses]

    # ----------------------------
    # Data collection
    # ----------------------------
    def Collect_Data(self, Tickers):
        """
        Collect all required Alpha Vantage datasets for each ticker and
        return a unified DataFrame of raw reports.
        """

        stock_info = self._fetch_json("OVERVIEW", Tickers)
        income_statement = self._fetch_json("INCOME_STATEMENT", Tickers)
        balance_sheet = self._fetch_json("BALANCE_SHEET", Tickers)
        cash_flow = self._fetch_json("CASH_FLOW", Tickers)
        insider_activity = self._fetch_json("INSIDER_TRANSACTIONS", Tickers)
        shares_outstanding = self._fetch_json("SHARES_OUTSTANDING", Tickers)

        return pd.DataFrame({
            'Ticker': Tickers,
            'Info': stock_info,
            'Income_Statement': income_statement,
            'Balance_Sheet': balance_sheet,
            'Cash_Flow': cash_flow,
            'Insider_Activity': insider_activity,
            'Shares_Outstanding': shares_outstanding
        })

    # ============================================================
    # ROIC Robustness Check
    # ============================================================
    def ROIC_CHK(self, InvestedCapital, ROIC, ROIC_AVG, Stock_Size, TotalAssets):
        """
        Determines whether ROIC is economically meaningful and stable enough
        to use for growth quality calculations.

        ROIC is rejected if:
        - Financial sector (capital structure distorts ROIC)
        - Missing or NaN values
        - Invested capital too small relative to assets
        - Capital volatility too high
        - Implausible average ROIC

        Output:
        - USE: list of boolean lists per stock/year
        """

        USE = []

        # Sector classification (used to exclude financials)
        Sector = [Overview[i].get('Sector') for i in range(len(Overview))]
        
        # Volatility of invested capital normalized by total assets
        VOL = [
            np.std(np.divide(ic, ta)[~np.isnan(np.divide(ic, ta))])
            if np.any(~np.isnan(np.divide(ic, ta)))
            else np.nan
            for ic, ta in zip(InvestedCapital, TotalAssets)
        ]

        for i in range(len(Stock_Size)):
            n_years = len(ROIC[i])
            years = range(1, n_years+1)
            use = []

            for j in years:
                epsilon = 0.05 * TotalAssets[i][j] if j < len(TotalAssets[i]) else 0
                roic_val = ROIC[i][j] if j < len(ROIC[i]) else np.nan

                if Sector[i] == 'FINANCIAL SERVICES':
                    use.append(False)
                elif roic_val is None or np.isnan(roic_val):
                    use.append(False)
                elif j >= len(TotalAssets[i]) or abs(InvestedCapital[i][j]) < epsilon:
                    use.append(False)
                elif np.isnan(VOL[i]) or VOL[i] > 0.35:
                    use.append(False)
                elif ROIC_AVG[i] < -0.2 or ROIC_AVG[i] > 0.6:
                    use.append(False)
                else:
                    use.append(True)

            USE.append(use)

        return USE
    
    def analysis_years(self, reports, max_years=5):
        """
        Returns a range of annual indices to analyze.

        - Skips index 0 (most recent year) intentionally
        - Limits analysis to `max_years`
        - Ensures we do not run past available data

        Example:
            reports = [2024, 2023, 2022, 2021, 2020]
            → returns range(1, 5)
        """
        n = len(reports)
        return range(1, min(n - 1, max_years + 1))
    
    def analysis_quarters(self, reports, max_years=5):
            """
            Determines number of quarters to analyze.
    
            - Used for quarterly data alignment
            - Caps analysis at max_years × 4 quarters
            - Returns 0 if reports are missing
            """
            if reports is None:
                return 0
            return min(len(reports), max_years * 4)
        
    def compute_CAGR(self, X):
        """
        Computes CAGR for a list of time series.

        Formula:
            CAGR = (Final / Initial)^(1 / periods) − 1

        Notes:
            - Assumes most recent value is at index 0
            - Requires at least 2 data points
            - Returns NaN if invalid or zero baseline
        """
        CAGR = []

        for x in X:
            periods = len(x) - 1
            if len(x) < 2 or x[-1] == 0 or np.isnan(x[-1]):
                CAGR.append(np.nan)
            else:
                cagr = (x[0] / x[-1]) ** (1 / periods) - 1
                CAGR.append(cagr)
    
        return CAGR

    def compute_EBIT(self, Stock_Size, IncomeStatement):
        """
        Extracts EBIT for each stock over analysis years.
        """
        EBIT = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)
                
            ebit = [
                self.safe_get_report(reports[j], 'ebit')
                for j in years
            ]
            
            EBIT.append(ebit)
            
        return EBIT

    def compute_EBITDA(self, Stock_Size, IncomeStatement):
        """
        Extracts EBITDA for each stock over analysis years.
        """
        EBITDA = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)
                
            ebitda = [
                self.safe_get_report(reports[j], 'ebitda')
                for j in years
            ]
            
            EBITDA.append(ebitda)
        
        return EBITDA

    def compute_TaxRate(self, Stock_Size, IncomeStatement):
        """
        Computes effective tax rate per year.

        Formula:
            incomeTaxExpense / incomeBeforeTax
        """
        TaxRate = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)
                
            tax_rate = [
                self.safe_get_report(reports[j], 'incomeTaxExpense') /
                self.safe_get_report(reports[j], 'incomeBeforeTax') if self.safe_get_report(reports[j], 'incomeBeforeTax') != 0 else np.nan
                for j in years
            ]
            
            TaxRate.append(tax_rate)
            
        return TaxRate

    def compute_NOPAT(self, EBIT, TaxRate):
        """
        Computes Net Operating Profit After Tax (NOPAT).

        Formula:
            NOPAT = EBIT × (1 − Tax Rate)
        """
        NOPAT = [
            list(np.array(e) * (1 - np.array(t)))
            for e, t in zip(EBIT, TaxRate)
        ]
        
        return NOPAT

    def compute_InvestedCapital(self, Stock_Size, BalSheet):
        """
        Computes Invested Capital.

        Formula:
            Equity + Short-Term Debt + Long-Term Debt + Capital Leases
        """
        IC = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            ic = []
            years = self.analysis_years(reports)

            for j in years:
                equity = self.safe_get_report(reports[j], 'totalShareholderEquity') or 0
                std = self.safe_get_report(reports[j], 'shortTermDebt') or 0
                ltd = self.safe_get_report(reports[j], 'longTermDebt') or 0
                leases = self.safe_get_report(reports[j], 'capitalLeaseObligations') or 0

                invested_capital = equity + std + ltd + leases
                ic.append(invested_capital)
                
            IC.append(ic)
        
        return IC

    def compute_Revenue(self, Stock_Size, IncomeStatement):
        """
        Extracts total revenue for CAGR analysis.
        """
        Revenue = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)
                
            revenue = [
                self.safe_get_report(reports[j], 'totalRevenue') or 0
                for j in years
            ]
            
            Revenue.append(revenue)
        
        return Revenue

    def compute_NetCapEx(self, Stock_Size, CashFlow):
        """
        Extracts capital expenditures from cash flow statements.
        """
        NetCapEx = []

        for i in range(len(Stock_Size)):
            reports = CashFlow[i]
            years = self.analysis_years(reports)
            
            netcapex = [
                self.safe_get_report(reports[j], 'capitalExpenditures')
                for j in years
            ]
            
            NetCapEx.append(netcapex)
        
        return NetCapEx

    def compute_Del_NWC(self, Stock_Size, BalSheet):
        """
        Computes change in Net Working Capital (ΔNWC).

        Formula:
            (CA − CL)_t+1 − (CA − CL)_t
        """
        Del_NWC = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            years = self.analysis_years(reports)
                
            del_nwc = [
                (
                    (self.safe_get_report(reports[j+1], 'totalCurrentAssets') -
                     self.safe_get_report(reports[j+1], 'totalCurrentLiabilities')) -
                    (self.safe_get_report(reports[j], 'totalCurrentAssets') -
                     self.safe_get_report(reports[j], 'totalCurrentLiabilities'))
                )
                for j in years
            ]
            
            Del_NWC.append(del_nwc)
            
        return Del_NWC

    def compute_FCF(self, NOPAT, DandA, Del_NWC, NetCapEx):
        """
        Computes Free Cash Flow.

        Formula:
            FCF = NOPAT + D&A − ΔNWC − CapEx
        """
        FCF = [
            list(
                np.array(nopat) +
                np.array(da) -
                np.array(delnwc) -
                abs(np.array(capex))
            )
            for nopat, da, delnwc, capex in zip(NOPAT, DandA, Del_NWC, NetCapEx)
        ]
    
        return FCF

    def compute_ROIC(self, Stock_Size, BalSheet, NOPAT, IC):
        """
        Computes Return on Invested Capital (ROIC).

        Formula:
            ROIC = NOPAT / Invested Capital
        """
        ROIC = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            roic_i = []
            years = self.analysis_years(reports)

            for j in years:
                nopat = NOPAT[i][j-1]
                invested_capital = IC[i][j-1]
                
                if invested_capital <= 0 or nopat is None:
                    roic_i.append(np.nan)
                else:
                    roic_i.append(nopat / invested_capital)

            ROIC.append(roic_i)
        
        return ROIC

    def compute_Total_Assets(self, Stock_Size, BalSheet):
        """
        Extracts total assets per year.
        """
        TotAssets = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            years = self.analysis_years(reports)
             
            ta = [
                self.safe_get_report(reports[j], 'totalAssets') or 0
                for j in years
            ]
            
            TotAssets.append(ta)
            
        return TotAssets

    def compute_Debt(self, Stock_Size, BalSheet):
        """
        Computes total debt = short-term + long-term debt.
        """
        Debt = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            years = self.analysis_years(reports)
                
            d = [
                (self.safe_get_report(reports[j], 'shortTermDebt') or 0) +
                (self.safe_get_report(reports[j], 'longTermDebt') or 0)
                for j in years
            ]
            
            Debt.append(d)
        
        return Debt

    def compute_DepreciationandAmortization(self, Stock_Size, IncomeStatement):
        """
        Extracts depreciation and amortization values.
        """
        DandA = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)
                
            da = [
                self.safe_get_report(reports[j], 'depreciationAndAmortization') or 0
                for j in years
            ]
            
            DandA.append(da)
        
        return DandA
    
    def compute_NetDebt(self, Stock_Size, BalSheet):
        """
        Computes Net Debt.

        Formula:
            Net Debt = Short-Term Debt + Long-Term Debt − Cash
        """
        Net_Debt = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            years = self.analysis_years(reports)
            
            nd = [
                (self.safe_get_report(reports[j], 'shortTermDebt') or 0) +
                (self.safe_get_report(reports[j], 'longTermDebt') or 0) -
                (self.safe_get_report(reports[j], 'cashAndCashEquivalentsAtCarryingValue') or 0)
                for j in years
            ]
            
            Net_Debt.append(nd)
            
        return Net_Debt

    def compute_liquidity_factor(self, BalSheet, IncomeStatement):
        # ------Liquidity------
        '''
        Liquidity measures the firm’s ability to meet short-term obligations
        without financial stress.

        Components:
            - Current Ratio = Current Assets / Current Liabilities
            - Quick Ratio = (Cash + Marketable Securities + Receivables) / Current Liabilities
            - Working Capital per Revenue = (Current Assets − Current Liabilities) / Revenue
        '''
        
        # -------------------------------------------------------
        # Current Ratio
        # -------------------------------------------------------
        # Measures raw short-term balance sheet coverage.
        # Divide-by-zero and invalid liability cases are explicitly guarded.

        current_ratio = []

        for stock in BalSheet:
            assets = self.safe_get_report(stock[0], 'totalCurrentAssets')
            liabilities = self.safe_get_report(stock[0], 'totalCurrentLiabilities')
            
            # Invalid or zero liabilities imply unusable liquidity signal
            if liabilities is None or liabilities <= 0 or not np.isfinite(liabilities):
                current_ratio.append(np.nan)
            else:
                current_ratio.append(assets / liabilities)
       
        # -------------------------------------------------------
        # Quick Ratio
        # -------------------------------------------------------
        # More conservative liquidity measure.
        # Excludes inventory and other less liquid current assets.

        quick_ratio = []

        for stock in BalSheet:
            # If current liabilities are missing or zero, liquidity cannot be assessed
            if self.safe_get_report(stock[0], 'totalCurrentLiabilities') in (None, 0): 
                quick_ratio.append(np.nan) 
            else: 
                quick_ratio.append(
                    (
                        (self.safe_get_report(stock[0], 'cashAndCashEquivalentsAtCarryingValue') or 0) +
                        (self.safe_get_report(stock[0], 'shortTermInvestments') or 0) +
                        (self.safe_get_report(stock[0], 'currentNetReceivables') or 0)
                    ) /
                    self.safe_get_report(stock[0], 'totalCurrentLiabilities')
                )
     
        # -------------------------------------------------------
        # Working Capital per Revenue
        # -------------------------------------------------------
        # Normalizes working capital by firm scale.
        # Penalizes companies that must carry excessive short-term capital
        # to support a given revenue base.

        Working_Cap_Per_Revenue = []

        for bs, is_ in zip(BalSheet, IncomeStatement):
            current_assets = self.safe_get_report(bs[0], 'totalCurrentAssets') or 0
            current_liabilities = self.safe_get_report(bs[0], 'totalCurrentLiabilities') or 0
            revenue = self.safe_get_report(is_[0], 'totalRevenue')
        
            # Revenue must exist for normalization to be meaningful
            if revenue in (None, 0):
                Working_Cap_Per_Revenue.append(np.nan)
            else:
                working_capital = current_assets - current_liabilities
                Working_Cap_Per_Revenue.append(working_capital / revenue)

        # -------------------------------------------------------
        # Standardization & Scoring
        # -------------------------------------------------------
        # Uses exponential saturation:
        #   SS = 100 × (1 − exp(−x))
        #
        # Properties:
        #   - Strongly penalizes weak liquidity (< 1)
        #   - Rewards improvements at low values
        #   - Models diminishing returns at high ratios
        #   - Keeps all scores bounded on [0, 100]

        # Standardized Current Ratio
        SS_current_ratio = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in current_ratio
                ]
        
        # Standardized Quick Ratio
        SS_quick_ratio = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in quick_ratio
                ]
        
        # Standardized Working Capital per Revenue
        SS_WCPR = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in Working_Cap_Per_Revenue
                ]
        
        # -------------------------------------------------------
        # Final Liquidity Score
        # -------------------------------------------------------
        # Equal-weighted aggregation of:
        #   - Balance sheet coverage (current ratio)
        #   - Liquid asset coverage (quick ratio)
        #   - Capital efficiency of operations (WCPR)

        SS_Total_Liquidity = np.mean(
            np.vstack([SS_current_ratio, SS_quick_ratio, SS_WCPR]),
            axis=0
        ).tolist()
        
        return SS_Total_Liquidity

    def compute_growth_factor(self, tickers, BalSheet, CashFlow, IncomeStatement):
        
        # ============================================================
        # GROWTH FACTOR
        # ============================================================
        """
        Growth Factor

        Components:
            1) Growth Quality
                - Reinvestment Rate × ROIC (preferred)
                - Falls back to Reinvestment Rate × ROE if ROIC is not robust
            2) Revenue CAGR (5y–10y or max available)
            3) Free Cash Flow CAGR (preferred over revenue)

        Final output:
            - Each sub-metric is standardized to a 0–100 score
            - Final Growth score is the mean of:
                [Growth Quality, Revenue CAGR, FCF CAGR]
        """

        # ------------------------------------------------------------
        # Determine stock size (number of balance sheet periods)
        # ------------------------------------------------------------
        Stock_Size = []

        for i in range(len(tickers)):
            stock_size = len(BalSheet[i])
            Stock_Size.append(stock_size)

        # ------------------------------------------------------------
        # Growth Quality = Reinvestment Rate × ROIC (or ROE fallback)
        # ------------------------------------------------------------

        # Net CapEx = CapEx - Depreciation
        NetCapEx = self.compute_NetCapEx(Stock_Size, CashFlow)

        # Change in Net Working Capital (ΔNWC)
        # Excludes most recent year to align timing with other metrics
        Del_NWC = self.compute_Del_NWC(Stock_Size, BalSheet)

        # EBIT from income statement
        EBIT = self.compute_EBIT(Stock_Size, IncomeStatement)

        # Effective tax rate
        TaxRate = self.compute_TaxRate(Stock_Size, IncomeStatement)

        # NOPAT = EBIT × (1 − Tax Rate)
        NOPAT = self.compute_NOPAT(EBIT, TaxRate)

        # Reinvestment Rate
        # (Net CapEx + ΔNWC) ÷ NOPAT
        ReInvRate = [
            list((np.array(nce) + np.array(d_nwc)) / np.array(nopat))
            for nce, d_nwc, nopat in zip(NetCapEx, Del_NWC, NOPAT)
        ]

        # Average reinvestment rate per stock
        ReInvRate_Avg = []
        for i in range(len(Stock_Size)):
            rir_avg = np.mean(np.array(ReInvRate[i]))
            ReInvRate_Avg.append(rir_avg)

        # ------------------------------------------------------------
        # ROIC Calculation
        # ROIC = NOPAT / Invested Capital
        # ------------------------------------------------------------

        IC = self.compute_InvestedCapital(Stock_Size, BalSheet)
        ROIC = self.compute_ROIC(Stock_Size, BalSheet, NOPAT, IC)

        # Average ROIC per stock (used for robustness checks)
        ROIC_AVG = [np.nanmean(roic) for i, roic in enumerate(ROIC)]

        # ------------------------------------------------------------
        # ROE Calculation (fallback if ROIC is unreliable)
        # ------------------------------------------------------------
        ROE = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            reports2 = BalSheet[i]
            years = self.analysis_years(reports)

            roe = []
            for j in years:
                ni = self.safe_get_report(reports[j], 'netIncome')
                eq = self.safe_get_report(reports2[j], 'totalShareholderEquity')

                if ni is None or eq is None or eq <= 0:
                    roe.append(np.nan)
                else:
                    roe.append(ni / eq)

            ROE.append(roe)

        # ------------------------------------------------------------
        # ROIC usability check
        # ------------------------------------------------------------

        # Total Assets used to evaluate ROIC stability and scale
        TotAssets = self.compute_Total_Assets(Stock_Size, BalSheet)

        # Boolean mask per year indicating whether ROIC is usable
        ROIC_Use = self.ROIC_CHK(IC, ROIC, ROIC_AVG, Stock_Size, TotAssets)

        # Ratio of usable ROIC years per stock
        Use_roic_ratio = []

        for i in range(len(Stock_Size)):
            reports = BalSheet[i]
            count = 0
            years = self.analysis_years(reports)

            for j in range(len(years)):
                if ROIC_Use[i][j] == True:
                    count += 1

            Use_roic_ratio.append(count / len(ROIC_Use[i]))

        # ------------------------------------------------------------
        # Final Growth Quality Calculation
        # ------------------------------------------------------------

        Growth_Quality = []

        # Use ROIC if robust ≥ 50% of years, else fall back to ROE
        for i in range(len(Use_roic_ratio)):
            if Use_roic_ratio[i] >= 0.5:
                gq = list(np.array(ReInvRate[i]) * np.array(ROIC[i]))
            else:
                gq = list(np.array(ReInvRate[i]) * np.array(ROE[i]))

            Growth_Quality.append(gq)

        # ------------------------------------------------------------
        # Revenue CAGR
        # ------------------------------------------------------------

        Revenue = self.compute_Revenue(Stock_Size, IncomeStatement)
        Revenue_CAGR = self.compute_CAGR(Revenue)

        # ------------------------------------------------------------
        # Free Cash Flow CAGR
        # ------------------------------------------------------------

        # Depreciation & Amortization from income statement
        DandA = []

        for i in range(len(Stock_Size)):
            reports = IncomeStatement[i]
            years = self.analysis_years(reports)

            da = [
                self.safe_get_report(reports[j], 'depreciationAndAmortization') or 0
                for j in years
            ]

            DandA.append(da)

        # FCF = NOPAT + D&A − ΔNWC − Net CapEx
        FCF = self.compute_FCF(NOPAT, DandA, Del_NWC, NetCapEx)

        # FCF CAGR
        FCF_CAGR = self.compute_CAGR(FCF)

        # ------------------------------------------------------------
        # Standardization (0–100)
        # ------------------------------------------------------------

        # Mean and volatility of Growth Quality
        GQ_mean = [
            np.nanmean(gq) if np.any(~np.isnan(gq)) else np.nan
            for gq in Growth_Quality
        ]

        GQ_SD = [
            np.nanstd(gq) if np.any(~np.isnan(gq)) else np.nan
            for gq in Growth_Quality
        ]

        # Kelly-style expected log growth
        GQ_Return = list(np.array(GQ_mean) - 0.5 * np.array(GQ_SD)**2)

        # Standardized Growth Quality score
        SS_GQ = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * (1 - np.exp(-i))
            for i in GQ_Return
        ]

        # Standardized Revenue CAGR score
        SS_Rev_CAGR = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * (1 - np.exp(-i))
            for i in Revenue_CAGR
        ]

        # Standardized FCF CAGR score
        SS_FCF_CAGR = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * (1 - np.exp(-i))
            for i in FCF_CAGR
        ]

        # ------------------------------------------------------------
        # Final Growth Factor
        # ------------------------------------------------------------

        # Equal-weight average of all growth subcomponents
        SS_Total_Growth = np.mean(
            np.vstack([SS_GQ, SS_Rev_CAGR, SS_FCF_CAGR]),
            axis=0
        ).tolist()

        return SS_Total_Growth

    def compute_capital_efficiency_factor(self, NOPAT, IC, FCF, Revenue, ROIC):
        
        #--------------Capital Efficiency-------------
        '''
        Capital Efficiency evaluates how effectively a company converts
        invested capital into economic value.

        Components:
            - Incremental ROIC (ROIIC): efficiency of *new* capital deployed
            - ROIC: baseline capital productivity
            - Free Cash Flow Margin: cash efficiency of operations

        Stability is explicitly rewarded via mean–variance penalization.
        '''

        # -------------------------------------------------------
        # Incremental ROIC (ROIIC)
        # -------------------------------------------------------
        # ROIIC = ΔNOPAT / ΔInvested Capital
        #
        # This captures whether *additional* capital is being deployed
        # at attractive returns, which is more informative than static ROIC.

        delNOPAT = [list(np.diff(nopat)) for nopat in NOPAT]
        delIC = [list(np.diff(ic)) for ic in IC]

        ROIIC = []

        for d_nopat, d_ic in zip(delNOPAT, delIC):
            d_nopat = np.array(d_nopat, dtype=float)
            d_ic = np.array(d_ic, dtype=float)

            # Avoid divide-by-zero and invalid IC changes
            roiic = np.where(
                (~np.isnan(d_ic)) & (d_ic != 0),
                d_nopat / d_ic,
                np.nan
            )

            ROIIC.append(roiic.tolist())

        # -------------------------------------------------------
        # Free Cash Flow Margin
        # -------------------------------------------------------
        # FCF Margin = Free Cash Flow / Revenue
        #
        # Measures how much cash the business produces per dollar of sales.
        # Rewards operational efficiency and pricing power.

        FCF_Margin = []

        for fcf, rev in zip(FCF, Revenue):
            fcf = np.array(fcf, dtype=float)
            rev = np.array(rev, dtype=float)

            fcfm = np.where(
                (~np.isnan(rev)) & (rev != 0),
                fcf / rev,
                np.nan
            )

            FCF_Margin.append(fcfm.tolist())

        # -------------------------------------------------------
        # ROIIC Stability & Expected Return
        # -------------------------------------------------------
        # Mean–variance framework inspired by Kelly Criterion:
        #
        # Expected Return ≈ mean − ½ × variance
        #
        # This rewards:
        #   - High capital efficiency
        #   - Consistency over time
        # And penalizes volatile or erratic capital deployment.

        ROIIC_mean = [
            np.nanmean(roiic) if np.any(~np.isnan(roiic)) else np.nan
            for roiic in ROIIC
        ]

        ROIIC_SD = [
            np.nanstd(roiic) if np.any(~np.isnan(roiic)) else np.nan
            for roiic in ROIIC
        ]

        ROIIC_Return = list(
            np.array(ROIIC_mean) - 0.5 * np.array(ROIIC_SD)**2
        )

        # -------------------------------------------------------
        # ROIC Stability & Expected Return
        # -------------------------------------------------------
        # Baseline capital productivity, evaluated over time.

        ROIC_mean = [
            np.nanmean(roic) if np.any(~np.isnan(roic)) else np.nan
            for roic in ROIC
        ]

        ROIC_SD = [
            np.nanstd(roic) if np.any(~np.isnan(roic)) else np.nan
            for roic in ROIC
        ]

        ROIC_Return = list(
            np.array(ROIC_mean) - 0.5 * np.array(ROIC_SD)**2
        )

        # -------------------------------------------------------
        # FCF Margin Stability & Expected Return
        # -------------------------------------------------------
        # Captures durability of cash conversion.
        # Penalizes businesses with lumpy or unreliable free cash flow.

        FCF_Margin_mean = [
            np.nanmean(fcfm) if np.any(~np.isnan(fcfm)) else np.nan
            for fcfm in FCF_Margin
        ]

        FCF_Margin_SD = [
            np.nanstd(fcfm) if np.any(~np.isnan(fcfm)) else np.nan
            for fcfm in FCF_Margin
        ]

        FCFM_Return = list(
            np.array(FCF_Margin_mean) - 0.5 * np.array(FCF_Margin_SD)**2
        )

        # -------------------------------------------------------
        # Standardization
        # -------------------------------------------------------
        # Exponential saturation:
        #   SS = 100 × (1 − exp(−x))
        #
        # Ensures:
        #   - Bad capital allocators score near zero
        #   - Good allocators are rewarded
        #   - Extreme outliers do not dominate the factor

        SS_ROIIC = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in ROIIC_Return
                ]

        SS_ROIC = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in ROIC_Return 
                ]

        SS_FCF_Margin = [
                0 if (i is None or np.isnan(i) or i <= 0)
                else 100 * (1 - np.exp(-i))
                for i in FCFM_Return 
                ]
        
        # -------------------------------------------------------
        # Final Capital Efficiency Score
        # -------------------------------------------------------
        # Equal-weighted blend of:
        #   - Incremental capital returns (ROIIC)
        #   - Baseline capital productivity (ROIC)
        #   - Cash generation efficiency (FCF Margin)

        SS_Total_CapEff = np.mean(
            np.vstack([SS_ROIIC, SS_ROIC, SS_FCF_Margin]),
            axis=0
        ).tolist()
        
        return SS_Total_CapEff

    def compute_coverage_factor(self, Stock_Size, Overview, BalSheet, IncomeStatement, FCF, EBIT):
        '''
        Coverage Factor
        Measures a company's ability to cover obligations such as debt and dividends.
        
        Components:
        - Dividend coverage: DPS / EPS (conditional, only for dividend-paying stocks)
        - Debt / EBITDA
        - FCF / Interest
        - EBIT / Interest (multi-year average + worst year)
        '''
    
        # -----------------------------
        # 1. Dividend Payout Ratio
        # -----------------------------
        # Get dividend per share (DPS) and earnings per share (EPS) for each stock from Overview
        DPS = [self.safe_get_report(report, 'DividendPerShare') or 0 for report in Overview]
        EPS = [self.safe_get_report(report, 'EPS') or 0 for report in Overview]
    
        Dividend_Payout = []
        for dps, eps in zip(DPS, EPS):
            # Validate values: skip invalid or non-positive EPS
            if dps is None or eps is None or eps <= 0 or not np.isfinite(dps) or not np.isfinite(eps):
                Dividend_Payout.append(np.nan)
            else:
                Dividend_Payout.append(dps / eps)  # Compute dividend payout ratio
    
        # -----------------------------
        # 2. Debt / EBITDA
        # -----------------------------
        # Debt = short-term debt + long-term debt
        Debt = self.compute_Debt(Stock_Size, BalSheet)
    
        # EBITDA from IncomeStatement
        EBITDA = self.compute_EBITDA(Stock_Size, IncomeStatement)
    
        D_E = []
        for d, ebitda in zip(Debt, EBITDA):
            d = np.array(d, dtype=float)
            ebitda = np.array(ebitda, dtype=float)
            valid = (~np.isnan(d)) & (~np.isnan(ebitda))  # Only valid entries
            d_e = np.full_like(ebitda, np.nan)           # Initialize with NaNs
            d_e[valid] = d[valid] / ebitda[valid]        # Debt / EBITDA
            D_E.append(d_e.tolist())
    
        # -----------------------------
        # 3. FCF / Interest
        # -----------------------------
        FCF_Interest = []
        for fcf, is_ in zip(FCF, IncomeStatement):
            series = []
            for t in range(len(fcf)):
                interest = self.safe_get_report(is_[t], 'interestExpense')
                if interest is None or interest <= 0:
                    series.append(np.nan)                  # Skip invalid interest
                else:
                    series.append(fcf[t] / interest)       # FCF / Interest
            FCF_Interest.append(series)
    
        # -----------------------------
        # 4. EBIT / Interest
        # -----------------------------
        EBIT_Interest = []
        for ebit, is_ in zip(EBIT, IncomeStatement):
            series = []
            for t in range(len(ebit)):
                interest = self.safe_get_report(is_[t], 'interestExpense')
                if interest is None or interest <= 0:
                    series.append(np.nan)
                else:
                    series.append(ebit[t] / interest)      # EBIT / Interest
            EBIT_Interest.append(series)
    
        # -----------------------------
        # 5. Standardization
        # -----------------------------
        # 5a. Debt / EBITDA standardization using log transformation (lower debt ratio → higher score)
        D_E_Return = []
        for de in D_E:
            de_arr = np.array(de, dtype=float)
            de_arr = de_arr[~np.isnan(de_arr)]           # Remove NaNs
            if len(de_arr) == 0:
                D_E_Return.append(np.nan)
            else:
                X = -np.log(de_arr)                     # Transform: lower D/E → higher score
                score = np.nanmean(X) - 0.5 * np.nanstd(X)**2  # Kelly-style expected return
                D_E_Return.append(score)
    
        # 5b. FCF / Interest standardization
        FCF_Interest_mean = [np.nanmean(fcfi) if np.any(~np.isnan(fcfi)) else np.nan for fcfi in FCF_Interest]
        FCF_Interest_SD = [np.nanstd(fcfi) if np.any(~np.isnan(fcfi)) else np.nan for fcfi in FCF_Interest]
        FCF_Interest_Return = list(np.array(FCF_Interest_mean) - 0.5 * np.array(FCF_Interest_SD)**2)
    
        # 5c. EBIT / Interest standardization
        EBIT_Interest_mean = [np.nanmean(ebiti) if np.any(~np.isnan(ebiti)) else np.nan for ebiti in EBIT_Interest]
        EBIT_Interest_SD = [np.nanstd(fcfi) if np.any(~np.isnan(fcfi)) else np.nan for fcfi in EBIT_Interest]
        EBIT_Interest_Return = list(np.array(EBIT_Interest_mean) - 0.5 * np.array(EBIT_Interest_SD)**2)
    
        # 5d. Dividend coverage standardization (capped at 200% payout)
        SS_Dividend_Coverage = []
        for payout in Dividend_Payout:
            if payout is None or np.isnan(payout) or payout <= 0:
                SS_Dividend_Coverage.append(0)
            else:
                capped = min(payout, 2.0)                  # cap at 200%
                SS_Dividend_Coverage.append(100 * np.exp(-capped))  # lower payout → higher score
    
        # 5e. Debt / EBITDA score
        SS_Debt_EBITDA = [0 if i is None or np.isnan(i) or i <= 0 else 100 * (1 - np.exp(-i)) for i in D_E_Return]
    
        # 5f. FCF / Interest score
        SS_FCF_Interest = [0 if i is None or np.isnan(i) or i <= 0 else 100 * (1 - np.exp(-i)) for i in FCF_Interest_Return]
    
        # 5g. EBIT / Interest score
        SS_EBIT_Interest = [0 if i is None or np.isnan(i) or i <= 0 else 100 * (1 - np.exp(-i)) for i in EBIT_Interest_Return]
    
        # -----------------------------
        # 6. Total Coverage Score
        # -----------------------------
        SS_Total_Cov = []
        for i in range(len(Stock_Size)):
            reports = Overview[i]
            # Determine if the stock pays dividends
            if self.safe_get_report(reports, 'DividendPerShare') == (0 or np.isnan(self.safe_get_report(reports, 'DividendPerShare'))):
                # Without dividend component
                ss_tot_cov = np.mean([SS_Debt_EBITDA[i], SS_FCF_Interest[i], SS_EBIT_Interest[i]])
            else:
                # With dividend component
                ss_tot_cov = np.mean([SS_Debt_EBITDA[i], SS_FCF_Interest[i], SS_EBIT_Interest[i], SS_Dividend_Coverage[i]])
            SS_Total_Cov.append(ss_tot_cov)
    
        return SS_Total_Cov

    def compute_value_factor(self, Stock_Size, Overview, BalSheet, EBIT):
        '''
        Value Factor
        Measures whether a stock is cheap or expensive relative to fundamentals.
    
        Components:
        - Graham Number (reference only)
        - Price-to-Book Ratio (P/B)
        - Price-to-Earnings Ratio (P/E)
        - Enterprise Value / EBIT
        '''
    
        # -----------------------------
        # 1. Graham Number (Reference)
        # -----------------------------
        # Graham Number formula: sqrt(22.5 * EPS * Book Value per Share)
        # Max price should not exceed 15x average earnings or 1.5x book value.
        BVPS = [self.safe_get_report(report, 'BookValue') or 0 for report in Overview]
        EPS = [self.safe_get_report(report, 'EPS') or 0 for report in Overview]
    
        Graham_Num = []
        for eps, bv in zip(EPS, BVPS):
            eps = np.array(eps, dtype=float)
            bv = np.array(bv, dtype=float)
    
            # Only valid positive EPS values
            valid = (~np.isnan(eps)) & (~np.isnan(bv)) & (eps > 0)
    
            gn = np.full_like(eps, np.nan)  # Initialize with NaN
    
            gn[valid] = np.sqrt(22.5 * eps[valid] * bv[valid])  # Compute Graham Number
    
            Graham_Num.append(gn.tolist())
    
        # -----------------------------
        # 2. Price-to-Book Ratio (P/B)
        # -----------------------------
        # Only meaningful for capital-intensive sectors
        PB = [self.safe_get_report(report, 'PriceToBookRatio') or 0 for report in Overview]
    
        # -----------------------------
        # 3. Price-to-Earnings Ratio (P/E)
        # -----------------------------
        PE = [self.safe_get_report(report, 'TrailingPE') or 0 for report in Overview]
    
        # -----------------------------
        # 4. Enterprise Value / EBIT
        # -----------------------------
        # Enterprise Value (EV) = Market Capitalization + Total Debt – Cash & Cash Equivalents
        MC = [self.safe_get_report(report, 'MarketCapitalization') or 0 for report in Overview]
        Total_Debt = self.compute_Debt(Stock_Size, BalSheet)
        CCE = [self.safe_get_report(report[0], 'cashAndCashEquivalentsAtCarryingValue') or 0 for report in BalSheet]
    
        # Calculate EV for each stock
        EV = [np.array(mc) + np.array(td[0]) - np.array(cce) for mc, td, cce in zip(MC, Total_Debt, CCE)]
    
        # EV / EBIT ratio
        EV_EBIT = [ev / ebit[0] for ev, ebit in zip(EV, EBIT)]
    
        # -----------------------------
        # 5. Standardization of Value Metrics
        # -----------------------------
        # Standardizing P/B: lower P/B = better (capped at 100)
        SS_PB = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else min(100, 100 / i)
            for i in PB
        ]
    
        # Standardizing P/E: lower P/E = better
        SS_PE = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else min(100, 100 / i)
            for i in PE
        ]
    
        # Standardizing EV/EBIT: lower EV/EBIT = better
        SS_EV_EBIT = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else min(100, 100 / i)
            for i in EV_EBIT
        ]
    
        # -----------------------------
        # 6. Combine All Value Scores
        # -----------------------------
        SS_Total_Value = np.mean(np.vstack([SS_PB, SS_PE, SS_EV_EBIT]), axis=0).tolist()
    
        return SS_Total_Value


    def compute_risk_factor(self, Stock_Size, Insiders, Overview, BalSheets, Revenue):
        '''
        Risk Factor
        Measures various risk dimensions of a stock.
    
        Components:
        - Insider Activity
            • Dollar volume / market cap
            • Share volume / shares outstanding
        - Dilution rate (over 5–10 years)
        - Revenue volatility (standard deviation / mean)
        - Net Debt / (Net Debt + Market Cap)
        '''
    
        # -----------------------------
        # 1. Insider Activity - Filter last year transactions
        # -----------------------------
        today = date.today()
        one_year_ago = today - timedelta(days=365)
        Date_Window = []
        
        insider_report = Insiders
        
        for i in range(len(Stock_Size)):
            
            raw_dates = self.safe_get_list(insider_report[i], 'data', 'transaction_date')
    
    
            parsed_dates = []
            for d in raw_dates:
                if d and d != "None":
                    try:
                        parsed_dates.append(datetime.strptime(d, "%Y-%m-%d").date())
                    except ValueError:
                        pass  # skip any invalid date strings
    
    
            recent_dates = [
            d for d in parsed_dates
            if one_year_ago <= d <= today
            ]
            
            Date_Window.append(recent_dates)
    
        # -----------------------------
        # 2. Insider Transaction Volumes
        # -----------------------------
        insider_report_shares_DW = []
        insider_report_price_DW = []
        insider_report_AD_DW = []
        insider_report_transact_DV = []
        insider_report_transact_SV = []
        Total_Insider_Dollar_Volume_YR = []
        Total_Insider_Share_Volume_YR = []
    
        for i in range(len(Stock_Size)):
            SR = []  # Shares
            SP = []  # Share prices
            AD = []  # Acquisition or Disposal
            TX_Dollar_Volume = []  # Transaction dollar volume
            TX_Share_Volume = []   # Transaction share volume
    
            shares_list = self.safe_get_list(insider_report[i], "data", "shares")
            price_list = self.safe_get_list(insider_report[i], "data", "share_price")
            AD_list = self.safe_get_list(insider_report[i], "data", "acquisition_or_disposal")
    
            for j in range(len(Date_Window[i])):
                shares_raw = shares_list[j]
                share_price = price_list[j]
                AorD = AD_list[j]
    
                # Skip missing or invalid data
                if shares_raw in [None, '', 'None'] or share_price in [None, '', 'None']:
                    continue
    
                shares_raw = float(shares_raw)
                share_price = float(share_price)
    
                # Make shares negative if disposal
                if AorD == "D":
                    shares_raw = -shares_raw
    
                txdv = shares_raw * share_price  # Dollar volume
                txsv = shares_raw               # Share volume
    
                SR.append(shares_raw)
                SP.append(share_price)
                AD.append(AorD)
                TX_Dollar_Volume.append(txdv)
                TX_Share_Volume.append(txsv)
    
            insider_report_shares_DW.append(SR)
            insider_report_price_DW.append(SP)
            insider_report_AD_DW.append(AD)
            insider_report_transact_DV.append(TX_Dollar_Volume)
            insider_report_transact_SV.append(TX_Share_Volume)
            Total_Insider_Dollar_Volume_YR.append(sum(TX_Dollar_Volume))
            Total_Insider_Share_Volume_YR.append(sum(TX_Share_Volume))
    
        # -----------------------------
        # 3. Market Capitalization & Shares Outstanding
        # -----------------------------
        MC = [self.safe_get_report(report, 'MarketCapitalization') or 0 for report in Overview]
        SO = [self.safe_get_report(report, 'SharesOutstanding') or 0 for report in Overview]
    
        # Ratios for insider activity
        DVtoMC = [np.divide(dv, mc) for dv, mc in zip(Total_Insider_Dollar_Volume_YR, MC)]
        SVtoSO = [np.divide(sv, so) for sv, so in zip(Total_Insider_Share_Volume_YR, SO)]
    
        # -----------------------------
        # 4. Dilution Rate over time
        # -----------------------------
        Basic_Shares = []
        Diluted_Shares = []
    
        for i in range(len(Stock_Size)):
            basic_shares = self.safe_get_list(Shares[i], "data", "shares_outstanding_basic")
            diluted_shares = self.safe_get_list(Shares[i], "data", "shares_outstanding_diluted")
    
            if not basic_shares or not diluted_shares:
                Basic_Shares.append([])
                Diluted_Shares.append([])
                continue
    
            n_quarters = self.analysis_quarters(basic_shares)
            Basic_Shares.append(basic_shares[:n_quarters])
            Diluted_Shares.append(diluted_shares[:n_quarters])
    
        Dilution_Rate = [
            1 - np.divide(np.array(bs[-1]).astype(float), np.array(ds[0]).astype(float))
            if bs and ds else 0  # or np.nan if you prefer
            for bs, ds in zip(Basic_Shares, Diluted_Shares)
        ]

    
        # -----------------------------
        # 5. Revenue Volatility
        # -----------------------------
        Rev_vol_norm = [
            np.std(rev) / np.mean(rev) if np.mean(rev) != 0 else 0
            for rev in Revenue
        ]
    
        # -----------------------------
        # 6. Net Debt / (Net Debt + Market Cap)
        # -----------------------------
        Net_Debt = self.compute_NetDebt(Stock_Size, BalSheets)
        NDtoCapital = [
            nd[0] / (nd[0] + mc) if (nd[0] + mc) != 0 else np.nan
            for nd, mc in zip(Net_Debt, MC)
        ]
        
        # -----------------------------
        # 7. Standardizing Each Risk Component
        # -----------------------------
        # Insider activity
        SS_DVtoMC = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * (1 - np.exp(-i))
            for i in DVtoMC
        ]
        SS_SVtoSO = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * (1 - np.exp(-i))
            for i in SVtoSO
        ]
    
        # Dilution
        SS_Dilution = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * np.exp(-i)
            for i in Dilution_Rate
        ]
    
        # Revenue volatility
        SS_Rev = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * np.exp(-i)
            for i in Rev_vol_norm
        ]
    
        # Net Debt / Capital
        SS_NDtoC = [
            0 if (i is None or np.isnan(i) or i <= 0)
            else 100 * np.exp(-i)
            for i in NDtoCapital
        ]
    
        # -----------------------------
        # 8. Aggregate Total Risk Score
        # -----------------------------
        SS_Total_Risk = np.mean(
            np.vstack([SS_DVtoMC, SS_SVtoSO, SS_Dilution, SS_Rev, SS_NDtoC]),
            axis=0
        ).tolist()
        
        return SS_Total_Risk

    def Factor_Model(self, Liquidity, Growth, CapEff, Coverage, Value, Risk, tickers):
        """
        Combines all standardized factor scores into a single Pandas DataFrame.
        
        Parameters:
        - Liquidity, Growth, CapEff, Coverage, Value, Risk: lists of standardized scores for each stock
        - tickers: list of ticker symbols corresponding to the scores
        
        Returns:
        - df: Pandas DataFrame with tickers as the index and factor scores as columns
        """
    
        # Create a DataFrame from the lists
        df = pd.DataFrame({
            'Ticker': tickers,    # Ticker symbols for each stock
            'Liquidity': Liquidity,
            'Growth': Growth,
            'CapEff': CapEff,
            'Coverage': Coverage,
            'Risk': Risk,
            'Value': Value
        })
    
        # Set the 'Ticker' column as the DataFrame index for easier access
        df.set_index('Ticker', inplace=True)
    
        return df
    
    def compute_Macro_Weight(self, Factors):
        
        #-------------- Macroeconomic Weights -------------
        '''
        Macroeconomic weighting calculates multiplicative weights applied to each
        company factor based on the current macroeconomic regime.
    
        Methodology:
        - Each macro variable is evaluated relative to its own 20-year historical
          distribution using a z-score.
        - Z-scores are converted into bounded tilts.
        - Tilts adjust factor importance depending on whether the macro factor
          has a DIRECT or INVERSE relationship with the company factor.
    
        Interpretation principle:
        - Macros describe the *environment companies operate in*, not company quality.
        - When macro conditions already favor a factor, that factor is UNDER-weighted.
        - When macro conditions are hostile, the factor is OVER-weighted to reward
          companies that outperform despite the environment.
        
        Factor ↔ Macro mapping:
            'Liquidity': No macro conditioning (Sharpe overlap, intentionally frozen)
            'Growth':    US GDP Growth                Source: https://data.worldbank.org/indicator/NY.GDP.MKTP.CD?end=2024&locations=US&start=2004
            'CapEff':    Global Commodities Index     Source: https://fred.stlouisfed.org/series/PALLFNFINDEXQ
            'Coverage':  CPI                          Source: https://www.bls.gov/charts/consumer-price-index/consumer-price-index-by-category-line-chart.htm
            'Risk':      S&P 500 Sharpe Ratio         Source: https://portfolioslab.com/tools/sharpe-ratio
            'Value':     Shiller CAPE                 Source: https://www.multpl.com/shiller-pe/table/by-year
        '''
   
        # Hard-coded annual macroeconomic data for the past ~20 years.
        # These are used to establish long-run means and standard deviations.
        # Sharpe ratio is handled separately due to lack of clean historical series.
        
        MACRO_DATA = {
        2006: {"CPI": 2.5, "Commodities": 121.30, "GDP": 2.784539637, "CAPE": 26.47},
        2007: {"CPI": 4.1, "Commodities": 137.66, "GDP": 2.003858298, "CAPE": 27.21},
        2008: {"CPI": 0.1, "Commodities": 173.57, "GDP": 0.11358725, "CAPE": 24.02},
        2009: {"CPI": 2.7, "Commodities": 119.03, "GDP": -2.57650023, "CAPE": 15.17},
        2010: {"CPI": 1.5, "Commodities": 144.68, "GDP": 2.695192577, "CAPE": 20.53},
        2011: {"CPI": 3.0, "Commodities": 179.55, "GDP": 1.564406858, "CAPE": 22.98},
        2012: {"CPI": 1.7, "Commodities": 172.55, "GDP": 2.289113387, "CAPE": 21.21},
        2013: {"CPI": 1.5, "Commodities": 166.05, "GDP": 2.117830098, "CAPE": 21.90},
        2014: {"CPI": 0.8, "Commodities": 158.10, "GDP": 2.523819814, "CAPE": 24.86},
        2015: {"CPI": 0.7, "Commodities": 107.92, "GDP": 2.945550455, "CAPE": 26.49},
        2016: {"CPI": 2.1, "Commodities": 100.00, "GDP": 1.819451478, "CAPE": 24.21},
        2017: {"CPI": 2.1, "Commodities": 113.33, "GDP": 2.457622299, "CAPE": 28.06},
        2018: {"CPI": 1.9, "Commodities": 127.73, "GDP": 2.966505069, "CAPE": 33.31},
        2019: {"CPI": 2.3, "Commodities": 117.02, "GDP": 2.583825334, "CAPE": 28.38},
        2020: {"CPI": 1.4, "Commodities": 105.86, "GDP": -2.16302914, "CAPE": 30.99},
        2021: {"CPI": 7.0, "Commodities": 161.50, "GDP": 6.055052932, "CAPE": 34.51},
        2022: {"CPI": 6.5, "Commodities": 215.92, "GDP": 2.51237532, "CAPE": 36.94},
        2023: {"CPI": 3.4, "Commodities": 165.65, "GDP": 2.887556007, "CAPE": 28.34},
        2024: {"CPI": 2.9, "Commodities": 164.82, "GDP": 2.793001277, "CAPE": 31.97},
        2025: {"CPI": 2.7, "Commodities": 166.90, "GDP": 8.28, "CAPE": 37.14},  # 2025 GDP YoY % proxy using Q3 2025 current-dollar annualized growth Source: https://www.jec.senate.gov/public/index.cfm/republicans/gdp-update 
    }
    
        # ---- Long-run macro statistics ----
        
        CPI_Mean = np.mean([MACRO_DATA[yr]["CPI"] for yr in MACRO_DATA])
        CPI_SD   = np.std([MACRO_DATA[yr]["CPI"] for yr in MACRO_DATA])
    
        GCI_Mean = np.mean([MACRO_DATA[yr]["Commodities"] for yr in MACRO_DATA])
        GCI_SD   = np.std([MACRO_DATA[yr]["Commodities"] for yr in MACRO_DATA])
    
        GDP_Mean = np.mean([MACRO_DATA[yr]["GDP"] for yr in MACRO_DATA])
        GDP_SD   = np.std([MACRO_DATA[yr]["GDP"] for yr in MACRO_DATA])
    
        CAPE_Mean = np.mean([MACRO_DATA[yr]["CAPE"] for yr in MACRO_DATA])
        CAPE_SD   = np.std([MACRO_DATA[yr]["CAPE"] for yr in MACRO_DATA])
    
        # ---- Current macro z-scores ----
        # Z-scores represent how extreme the current environment is relative to history.
        
        CPI_Z_Cur   = self.calculate_z(MACRO_DATA[2025]["CPI"], CPI_Mean, CPI_SD)
        GCI_Z_Cur   = self.calculate_z(MACRO_DATA[2025]["Commodities"], GCI_Mean, GCI_SD)
        GDP_Z_Cur   = self.calculate_z(MACRO_DATA[2025]["GDP"], GDP_Mean, GDP_SD)
        CAPE_Z_Cur  = self.calculate_z(MACRO_DATA[2025]["CAPE"], CAPE_Mean, CAPE_SD)
    
        # ---- Sharpe ratio handling ----
        # No long historical series available, so we infer mean and SD from
        # known distribution percentiles of portfolio Sharpe ratios.
        #
        # Portfolios:
        #   Median: 1.16
        #   75th percentile: 1.69
        #   99th percentile: 4.12
        
        Sharpe_Mean = 1.16
        Sharpe_SD   = (1.69 - 1.16) / 0.67  # 0.67 ≈ z-score of 75th percentile
        Sharpe_Current = 0.75               # Current S&P 500 Sharpe (approx.)
    
        Sharpe_Z_Cur = self.calculate_z(Sharpe_Current, Sharpe_Mean, Sharpe_SD)
    
        # ---- Macro → factor weighting logic ----
        #
        # DIRECT relationship:
        #   +z → macro environment already favorable → UNDER-weight the factor
        #
        # INVERSE relationship:
        #   +z → macro environment hostile → OVER-weight the factor
        #
        # This rewards companies that perform well despite macro headwinds.
        
        macro_weights = {
            "Liquidity": 1.0,  # intentionally not macro-conditioned (Sharpe overlap)
            "Growth":   self.multiplier(GDP_Z_Cur,   'direct'),   # high GDP → suppress growth
            "CapEff": self.multiplier(GCI_Z_Cur,   'inverse'),  # high CPI → emphasize coverage
            "Coverage":   self.multiplier(CPI_Z_Cur,   'inverse'),  # high commodities → emphasize efficiency
            "Risk":     self.multiplier(Sharpe_Z_Cur,'direct'),   # high Sharpe → suppress risk factor
            "Value":    self.multiplier(CAPE_Z_Cur,  'inverse')   # high CAPE → emphasize value
        }
        
        Factors_Macro_Adjusted = Factors.mul(pd.Series(macro_weights))
        
        return macro_weights, Factors_Macro_Adjusted
    
    def Spider_Visualize(self, Factors):
        """
        Creates a radar (spider) chart to visualize factor scores for multiple stocks.
    
        Parameters:
        - Factors: DataFrame where:
            - rows = stocks (index = ticker symbols)
            - columns = standardized factor scores ['Liquidity','Growth','CapEff','Coverage','Risk','Value']
    
        Visualization:
        - Each stock is a separate line on the radar chart
        - Filled area under each line for visual emphasis
        """
    
        # Extract factor labels and the number of variables
        labels = Factors.columns
        num_vars = len(labels)
    
        # Compute angles for each axis on the radar chart
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
        angles += angles[:1]  # close the loop so the polygon is complete
    
        # Initialize polar plot
        fig, ax = plt.subplots(figsize=(6,6), subplot_kw=dict(polar=True))
    
        # Plot each stock's scores
        for ticker in Factors.index:
            values = Factors.loc[ticker].tolist()  # get factor scores for this stock
            values += values[:1]  # close the loop
            ax.plot(angles, values, label=ticker, linewidth=2)  # line for stock
            ax.fill(angles, values, alpha=0.25)  # shaded area under the line
    
        # Adjust chart orientation
        ax.set_theta_offset(np.pi / 2)  # start from top
        ax.set_theta_direction(-1)      # clockwise direction
    
        # Set labels on the axes
        ax.set_thetagrids(np.degrees(angles[:-1]), labels)
    
        # Set range of values
        ax.set_ylim(0, 100)  # all factor scores normalized to 0-100
    
        # Add title and legend
        ax.set_title("Stock Factor Comparison", y=1.1)
        ax.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
    
        # Display the chart
        plt.show()
    
        return None
    
    def export_stock_factors_to_excel(self, normal_df, macro_df, output_file="stock_factors_analysis.xlsx"):
        """ 
        Exports both normal and macro-adjusted factor DataFrames to Excel with:
            - Composite score
            - Suggested portfolio allocation (%)
            - Delta between Normal and Macro-adjusted
            - Discrete color-coded factor values: Green >=70, Yellow 31-69, Red 0-30
            - Gradient color for Composite (green=high, red=low)
    
        Parameters:
            normal_df (pd.DataFrame): Index=Ticker, Columns=['Liquidity', 'Growth', 'CapEff', 'Coverage', 'Value', 'Risk']
            macro_df (pd.DataFrame): Same shape as normal_df
            output_file (str): Path to save Excel file
        """
    
        factor_cols = ['Liquidity', 'Growth', 'CapEff', 'Coverage', 'Value', 'Risk']
    
        # ---- Prepare Normal and Macro DataFrames ----
        def prepare_df(df):
            df = df.copy()
            df['Composite'] = df[factor_cols].sum(axis=1)
            avg_comp = df['Composite'].mean()
            df['Suggested Allocation (%)'] = df['Composite'] / avg_comp * (100 / len(df))
            return df
    
        normal_df = prepare_df(normal_df)
        macro_df = prepare_df(macro_df)
    
        # Delta
        delta_df = macro_df - normal_df
        delta_df['Composite'] = macro_df['Composite'] - normal_df['Composite']
        delta_df['Suggested Allocation (%)'] = macro_df['Suggested Allocation (%)'] - normal_df['Suggested Allocation (%)']
    
        # ---- Write to Excel ----
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            normal_df.to_excel(writer, sheet_name="Factors_Normal")
            macro_df.to_excel(writer, sheet_name="Factors_Macro")
            delta_df.to_excel(writer, sheet_name="Delta")
    
        wb = load_workbook(output_file)
    
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
    
        # ---- Function to apply coloring and borders ----
        def style_sheet(sheet_name):
            ws = wb[sheet_name]
            n_rows = ws.max_row - 1
            n_cols = len(factor_cols)
    
            # Factor coloring
            for col_idx, col_name in enumerate(factor_cols, start=2):
                for row_idx in range(2, 2 + n_rows):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(cell.value)
                    except (TypeError, ValueError):
                        continue
    
                    if val >= 70:
                        cell.fill = green_fill
                    elif 31 <= val <= 69:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill
    
                    cell.border = thin_border
    
            # Gradient coloring for Composite
            comp_col_idx = n_cols + 2
            comp_vals = []
            for r in range(2, 2 + n_rows):
                try:
                    comp_vals.append(float(ws.cell(row=r, column=comp_col_idx).value))
                except (TypeError, ValueError):
                    comp_vals.append(0)
    
            min_comp, max_comp = min(comp_vals), max(comp_vals)
            for row_idx, val in enumerate(comp_vals, start=2):
                if max_comp - min_comp == 0:
                    norm = 0.5
                else:
                    norm = (val - min_comp) / (max_comp - min_comp)
                red = int((1 - norm) * 255)
                green = int(norm * 255)
                fill = PatternFill(start_color=f"{red:02X}{green:02X}00",
                                   end_color=f"{red:02X}{green:02X}00", fill_type="solid")
                ws.cell(row=row_idx, column=comp_col_idx).fill = fill
                ws.cell(row=row_idx, column=comp_col_idx).border = thin_border
    
            # Borders for Suggested Allocation column
            alloc_col_idx = comp_col_idx + 1
            for row_idx in range(2, 2 + n_rows):
                ws.cell(row=row_idx, column=alloc_col_idx).border = thin_border
    
        # ---- Apply styling to all three sheets ----
        for sheet in ["Factors_Normal", "Factors_Macro", "Delta"]:
            style_sheet(sheet)
    
        wb.save(output_file)
        print(f"Excel file saved at {output_file} with Normal, Macro, and Delta sections")
